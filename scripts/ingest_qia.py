# Run as: PYTHONPATH=. python scripts/ingest_qia.py --file /path/to/QIA_file.xlsx
#         PYTHONPATH=. python scripts/ingest_qia.py --file /path/to/QIA_2026-01.xls
#
# ingest_qia.py — QIA quarantine data ingestion for Velvet Knowledge Hub.
#
# Supports two file types via auto-detection:
#
#   Type A — Annual XLSX (2019-2025): QIA_20*_ANNUAL_LIVESTOCK_QUARANTINE.xlsx
#     - True binary XLSX, parsed with openpyxl.
#     - Wide-format: months as column pairs (건수, 수량) for each country row.
#     - Sheet used: the sheet whose name contains "수입축산물".
#     - Filters to deer velvet products: 녹용, 생녹용 (excludes 녹각 — hard antler).
#     - Date format: "YYYY-MM" (melted from the monthly column headers).
#
#   Type B — Monthly HTML-XLS (2026): QIA_2026-0*_VELVET_IMPORT.xls
#     - HTML document masquerading as XLS (Content-Type: application/vnd.ms-excel, UTF-8).
#     - Detected by: extension == .xls AND first non-whitespace byte is '<'.
#     - Parsed with html.parser (stdlib).
#     - Columns: 국 가, 품 명, 건 수, 중량(kg/ea).
#     - Skips: 국가계 rows (subtotals), 총 계 rows (grand totals).
#     - Date format: "YYYY-MM" extracted from filename.
#
# Both types upsert into the VTW_Trade_Monthly tab with series = "korea_quarantine".
#
# Usage:
#   PYTHONPATH=. python scripts/ingest_qia.py --file /path/to/QIA_2024.xlsx
#   PYTHONPATH=. python scripts/ingest_qia.py --file /path/to/QIA_2026-01.xls
#   PYTHONPATH=. python scripts/ingest_qia.py --file /path/to/QIA_2026-01.xls --dry-run
#
# L-1:  PYTHONPATH=. ensures repo root is importable.
# L-2:  .env must be at repo root (/Users/Qs/C/velvet-knowledge-hub/.env).
# L-3:  GOOGLE_SERVICE_ACCOUNT_JSON must be single-line JSON in .env.
# L-4:  get_all_records() called once; new rows written in one append_rows() call.
# L-9:  hs_code stored as TEXT dot notation ("0507.90") — not cast to int.
# L-10: Dedup key is (date, series, hs_code, unit, country) — five fields.
# L-13: Columns detected by header name scan, never by fixed index.
#
# Security: no credentials or secrets in this file. All secrets from .env only.

import argparse
import json
import logging
import os
import re
import sys
import time
from html.parser import HTMLParser
from pathlib import Path

import gspread
import openpyxl
import yaml
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials

# ---------------------------------------------------------------------------
# L-1: ensure repo root is on sys.path.
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

logging.basicConfig(
    level=logging.WARNING,
    format="%(levelname)s %(name)s — %(message)s",
)
logger = logging.getLogger("ingest_qia")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = REPO_ROOT / "config.yaml"
TARGET_TAB = "VTW_Trade_Monthly"
SERIES_VALUE = "korea_quarantine"
SERIES_VALUE_EXPORT = "korea_quarantine_export"

# Sheets API only — no Drive API required (L-5 workaround).
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# HS code to assign to all QIA quarantine rows (dot notation TEXT — VKH schema).
HS_CODE = "0507.90"
HS_LABEL = "Horns, antlers, hooves, nails, claws and beaks — other"

# Units produced by this source.
UNIT_SHIPMENTS = "shipments"
UNIT_KG = "KG"

# Deer velvet product names to include (Korean). 녹각 = hard antler, excluded.
_VELVET_PRODUCTS = {"녹용", "생녹용"}

# Summary row markers — skip rows whose 품명 or 국가명 starts with these.
_SKIP_PREFIXES = ("국가계", "품명계", "총 계", "총계", "합 계", "합계", "계")

# L-4: batch write sleep between gspread calls.
_BATCH_SLEEP = 1.1


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _safe_str(value) -> str:
    """Return stripped string, or empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(raw) -> float:
    """Parse a number cell (int, float, or comma-formatted string) to float."""
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    cleaned = str(raw).strip().replace(",", "").replace("-", "0")
    if not cleaned:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _extract_ym_from_filename(filename: str) -> str:
    """
    Extract YYYY-MM from filename if present (monthly files).
    Falls back to year-only string (annual files).
    Returns empty string if no year found.
    """
    # Try YYYY-MM first (monthly: QIA_2026-01_VELVET_IMPORT.xls).
    match = re.search(r"(20\d{2}-\d{2})", filename)
    if match:
        return match.group(1)
    # Fall back to year only (annual: QIA_2024_ANNUAL_LIVESTOCK_QUARANTINE.xlsx).
    match = re.search(r"(20\d{2})", filename)
    if match:
        return match.group(1)
    return ""


# ---------------------------------------------------------------------------
# Type A — annual XLSX parser
# ---------------------------------------------------------------------------

def _find_import_sheet(wb) -> object:
    """
    Return the worksheet that contains '수입축산물' in its name.
    Raises ValueError if not found.
    """
    for sheet_name in wb.sheetnames:
        if "수입축산물" in sheet_name:
            return wb[sheet_name]
    raise ValueError(
        f"No sheet containing '수입축산물' found. Available sheets: {wb.sheetnames}"
    )


def _detect_annual_header(all_rows: list) -> tuple[int, int, int, int, list[tuple[int, str]]]:
    """
    Scan the first 10 rows to find the two-row header structure.

    Returns:
        (header_row_idx, col_product_category, col_product_name, col_country,
         month_columns)

    where month_columns is a list of (col_index, "YYYY-MM") pairs for each month.

    L-13: positions detected dynamically by content, never by fixed index.
    """
    for row_idx in range(min(10, len(all_rows) - 1)):
        row = [_safe_str(c) for c in all_rows[row_idx]]
        next_row = [_safe_str(c) for c in all_rows[row_idx + 1]]

        # Look for a row containing 품명 and 국가명 (or 국 가 or 국가).
        # Normalise by removing internal spaces before matching (handles "품  명" in 2021 files).
        has_product = any("품명" in c.replace(" ", "") for c in row)
        has_country = any("국가" in c.replace(" ", "") for c in row)

        if not (has_product and has_country):
            continue

        # Found the header row. Locate column indices.
        col_product_category = -1
        col_product_name = -1
        col_country = -1

        for col_idx, cell in enumerate(row):
            # 품목명 = category (e.g. 육류, 기타축산물)
            if "품목명" in cell or "품 목" in cell:
                if col_product_category == -1:
                    col_product_category = col_idx
            # 품명 = product name (e.g. 녹용) — normalise spaces ("품  명" in 2021)
            if cell.replace(" ", "") == "품명":
                if col_product_name == -1:
                    col_product_name = col_idx
            # 국가명 / 국 가 / 국가
            if "국가" in cell:
                if col_country == -1:
                    col_country = col_idx

        if col_product_name == -1 or col_country == -1:
            continue

        # Detect month columns: look for cells matching YYYY.MM pattern.
        month_columns: list[tuple[int, str]] = []
        year_month_re = re.compile(r"(20\d{2})\.\d{2}")

        # Month year labels are in this header row; sub-metric in the next row.
        # Each month occupies two columns: (건수, 수량).
        for col_idx, cell in enumerate(row):
            m = year_month_re.match(cell)
            if m:
                # Convert "2024.01" to "2024-01".
                ym = cell.replace(".", "-")
                month_columns.append((col_idx, ym))

        if not month_columns:
            continue

        logger.info(
            "Annual XLSX header row %d: category=%d, product=%d, country=%d, months=%d",
            row_idx, col_product_category, col_product_name, col_country, len(month_columns),
        )
        return row_idx, col_product_category, col_product_name, col_country, month_columns

    raise ValueError(
        "Could not detect annual XLSX header row. "
        "Expected a row containing '품명' and '국가명' with YYYY.MM month columns."
    )


def parse_qia_annual_xlsx(filepath: Path) -> list[dict]:
    """
    Parse a QIA annual livestock quarantine XLSX file.

    Filters to deer velvet products (녹용, 생녹용). Melts monthly columns into
    long-format rows: one shipments row + one KG row per (month, product, country).

    Returns VTW_Trade_Monthly schema rows.

    L-9:  hs_code TEXT dot notation.
    L-13: column positions detected by header scan.
    """
    logger.info("Parsing QIA annual XLSX: %s", filepath)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = _find_import_sheet(wb)

    all_rows = list(ws.iter_rows(values_only=True))

    (
        header_row_idx,
        col_category,
        col_product,
        col_country,
        month_columns,
    ) = _detect_annual_header(all_rows)

    # Each month has two sub-columns: col_idx = 건수, col_idx+1 = 수량(Kg,Ea).
    data_rows = all_rows[header_row_idx + 2:]  # Skip header row + sub-header row.

    output: list[dict] = []
    current_category = ""
    current_product = ""

    for row in data_rows:
        # Skip completely empty rows.
        if all(v is None or _safe_str(v) == "" for v in row):
            continue

        row_strs = [_safe_str(c) if i < len(row) else "" for i, c in enumerate(row)]

        # Forward-fill category and product name (QIA uses merged cells).
        category_cell = row_strs[col_category] if col_category != -1 and col_category < len(row_strs) else ""
        product_cell = row_strs[col_product] if col_product < len(row_strs) else ""
        country_cell = row_strs[col_country] if col_country < len(row_strs) else ""

        if category_cell:
            current_category = category_cell
        if product_cell:
            current_product = product_cell

        # Skip summary rows (국가계, 품명계, 총계, etc.).
        if any(country_cell.startswith(p) for p in _SKIP_PREFIXES):
            continue
        if any(current_product.startswith(p) for p in _SKIP_PREFIXES):
            continue

        if not country_cell or not current_product:
            continue

        # Filter to deer velvet products only.
        if current_product not in _VELVET_PRODUCTS:
            continue

        notes = f"source: {filepath.name} | product: {current_product}"

        for col_idx, ym in month_columns:
            # 건수 is at col_idx; 수량 is at col_idx + 1.
            count_raw = row[col_idx] if col_idx < len(row) else None
            weight_raw = row[col_idx + 1] if (col_idx + 1) < len(row) else None

            count_val = _parse_number(count_raw)
            weight_val = _parse_number(weight_raw)

            # Skip months with no activity.
            if count_val == 0 and weight_val == 0:
                continue

            # Shipments row.
            output.append({
                "date":     ym,
                "series":   SERIES_VALUE,
                "hs_code":  HS_CODE,
                "hs_label": HS_LABEL,
                "value":    count_val,
                "unit":     UNIT_SHIPMENTS,
                "country":  country_cell,
                "notes":    notes,
            })

            # KG row.
            output.append({
                "date":     ym,
                "series":   SERIES_VALUE,
                "hs_code":  HS_CODE,
                "hs_label": HS_LABEL,
                "value":    weight_val,
                "unit":     UNIT_KG,
                "country":  country_cell,
                "notes":    notes,
            })

    logger.info(
        "QIA annual parse complete: %d long-format records from %s.",
        len(output), filepath.name,
    )
    return output


# ---------------------------------------------------------------------------
# Type B — monthly HTML-XLS parser
# ---------------------------------------------------------------------------

class _TableParser(HTMLParser):
    """
    Minimal HTML table parser for QIA monthly XLS files.

    Captures ALL nested tables as separate entries in self.tables.
    Uses a stack to handle nested <table> elements correctly.

    Each entry in self.tables is a list of rows (list of cell strings).
    The data table is identified by _find_data_table after parsing.
    """

    def __init__(self):
        super().__init__()
        self.tables: list[list[list[str]]] = []
        # Stack of (table_rows, current_row, current_cell, in_cell) tuples.
        self._stack: list[dict] = []

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._stack.append({
                "rows": [],
                "current_row": [],
                "current_cell": "",
                "in_cell": False,
            })
        elif tag == "tr" and self._stack:
            self._stack[-1]["current_row"] = []
        elif tag in ("td", "th") and self._stack:
            self._stack[-1]["in_cell"] = True
            self._stack[-1]["current_cell"] = ""

    def handle_endtag(self, tag):
        if tag == "table" and self._stack:
            frame = self._stack.pop()
            if frame["rows"]:
                self.tables.append(frame["rows"])
        elif tag == "tr" and self._stack:
            frame = self._stack[-1]
            if frame["current_row"]:
                frame["rows"].append(frame["current_row"])
            frame["current_row"] = []
        elif tag in ("td", "th") and self._stack:
            frame = self._stack[-1]
            if frame["in_cell"]:
                frame["current_row"].append(frame["current_cell"].strip())
                frame["in_cell"] = False
                frame["current_cell"] = ""

    def handle_data(self, data):
        if self._stack and self._stack[-1]["in_cell"]:
            self._stack[-1]["current_cell"] += data


def _find_data_table(tables: list[list[list[str]]]) -> list[list[str]]:
    """
    Return the table that contains the quarantine data.

    The data table has a header row with ALL of: '국 가', '품 명', '건 수'.
    Requires at least 4 columns in the header row to avoid matching title rows.
    Returns the first table that matches; raises ValueError if not found.
    """
    for table in tables:
        for row in table[:5]:
            if len(row) < 4:
                continue
            row_joined = " ".join(row)
            # Must have column headers for country, product name, and count.
            has_country = "국 가" in row_joined or any(
                c.replace(" ", "") == "국가" for c in row
            )
            has_product = "품 명" in row_joined or any(
                c.replace(" ", "") == "품명" for c in row
            )
            has_count = "건 수" in row_joined or any(
                c.replace(" ", "") == "건수" for c in row
            )
            if has_country and has_product and has_count:
                return table
    raise ValueError(
        "No data table with '국 가', '품 명', '건 수' columns found in HTML-XLS file. "
        "Check the file structure."
    )


def _detect_html_columns(header_row: list[str]) -> dict[str, int]:
    """
    Map column names to indices for a QIA monthly HTML-XLS header row.

    Expected headers (with spaces): '국 가', '품 명', '건 수', '중량(kg/ea)'.
    L-13: detected by name, never by fixed index.
    """
    col_map: dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        normalised = cell.replace(" ", "").lower()
        if "국가" in normalised and "country" not in col_map:
            col_map["country"] = col_idx
        elif "품명" in normalised and "product" not in col_map:
            col_map["product"] = col_idx
        elif "건수" in normalised and "count" not in col_map:
            col_map["count"] = col_idx
        elif "중량" in normalised and "weight" not in col_map:
            col_map["weight"] = col_idx
    return col_map


def parse_qia_monthly_html_xls(filepath: Path) -> list[dict]:
    """
    Parse a QIA monthly HTML-XLS file into VTW_Trade_Monthly schema rows.

    The file is a UTF-8 HTML document with a .xls extension. It must be
    detected as HTML before calling this function (first byte is '<').

    Each data row emits TWO output rows:
      1. unit = "shipments", value = 건 수
      2. unit = "KG",        value = 중량(kg/ea)

    Skip rules:
      - Rows where 품 명 starts with "국가계", "총 계", etc. (subtotals).
      - Rows where both count == 0 AND weight == 0.

    Date: extracted as YYYY-MM from filename.
    L-13: column positions detected by header scan.
    """
    logger.info("Parsing QIA monthly HTML-XLS: %s", filepath)

    ym = _extract_ym_from_filename(filepath.name)
    if not re.match(r"20\d{2}-\d{2}", ym):
        logger.warning(
            "Could not extract YYYY-MM from filename %r — date will be empty.",
            filepath.name,
        )

    content = filepath.read_text(encoding="utf-8", errors="replace")

    parser = _TableParser()
    parser.feed(content)

    if not parser.tables:
        raise ValueError(f"No HTML tables found in {filepath.name}")

    data_table = _find_data_table(parser.tables)

    # Find header row (first row containing '국 가').
    header_row_idx = None
    for row_idx, row in enumerate(data_table):
        row_joined = " ".join(row)
        if "국 가" in row_joined or "국가" in row_joined:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        raise ValueError(f"Header row not found in {filepath.name}")

    header_row = data_table[header_row_idx]
    col_map = _detect_html_columns(header_row)

    required = {"country", "product", "count", "weight"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Missing columns {missing} in {filepath.name}. "
            f"Header row: {header_row}"
        )

    output: list[dict] = []
    current_country = ""
    notes = f"source: {filepath.name}"

    for row in data_table[header_row_idx + 1:]:
        if not row or all(c == "" for c in row):
            continue

        country_cell = row[col_map["country"]] if col_map["country"] < len(row) else ""
        product_cell = row[col_map["product"]] if col_map["product"] < len(row) else ""

        # Forward-fill country (HTML table uses rowspan).
        if country_cell:
            current_country = country_cell

        # Skip summary rows.
        if any(product_cell.startswith(p) for p in _SKIP_PREFIXES):
            continue
        if any(current_country.startswith(p) for p in _SKIP_PREFIXES):
            continue

        if not current_country or not product_cell:
            continue

        count_val = _parse_number(row[col_map["count"]] if col_map["count"] < len(row) else None)
        weight_val = _parse_number(row[col_map["weight"]] if col_map["weight"] < len(row) else None)

        if count_val == 0 and weight_val == 0:
            continue

        # Shipments row.
        output.append({
            "date":     ym,
            "series":   SERIES_VALUE,
            "hs_code":  HS_CODE,
            "hs_label": HS_LABEL,
            "value":    count_val,
            "unit":     UNIT_SHIPMENTS,
            "country":  current_country,
            "notes":    notes,
        })

        # KG row.
        output.append({
            "date":     ym,
            "series":   SERIES_VALUE,
            "hs_code":  HS_CODE,
            "hs_label": HS_LABEL,
            "value":    weight_val,
            "unit":     UNIT_KG,
            "country":  current_country,
            "notes":    notes,
        })

    logger.info(
        "QIA monthly HTML-XLS parse complete: %d records from %s.",
        len(output), filepath.name,
    )
    return output


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def _is_html_xls(filepath: Path) -> bool:
    """
    Return True if filepath is a HTML-disguised XLS file.

    Detection: extension == .xls AND first non-whitespace byte is '<'.
    """
    if filepath.suffix.lower() != ".xls":
        return False
    try:
        with filepath.open("rb") as fh:
            # Read first 64 bytes — QIA monthly files begin with up to 19 bytes
            # of whitespace (LF/CR) before the opening '<' HTML tag.
            head = fh.read(64).lstrip()
        return head.startswith(b"<")
    except OSError:
        return False


def _derive_series(filepath: Path) -> str:
    """
    Derive the series value from the filename.

    Files containing 'EXPORT' (case-insensitive) → 'korea_quarantine_export'.
    All other files (IMPORT, FRESH_IMPORT, annual XLSX) → 'korea_quarantine'.

    GAP-2 fix (2026-05-30): import and export series are kept separate so that
    export rows have distinct dedup keys from import rows and both are retained.
    The existing 'korea_quarantine' series is not renamed — backward-compatible.
    """
    if "EXPORT" in filepath.name.upper():
        return SERIES_VALUE_EXPORT
    return SERIES_VALUE


def parse_qia_file(filepath: Path) -> list[dict]:
    """
    Auto-detect file format and dispatch to the correct parser.

    .xlsx → openpyxl annual XLSX parser.
    .xls + HTML content → HTML table parser (monthly 2026 format).

    Series is derived from the filename: EXPORT files → 'korea_quarantine_export';
    all others → 'korea_quarantine'. GAP-2 fix (2026-05-30).
    """
    series = _derive_series(filepath)

    if filepath.suffix.lower() == ".xlsx":
        rows = parse_qia_annual_xlsx(filepath)
    elif _is_html_xls(filepath):
        rows = parse_qia_monthly_html_xls(filepath)
    else:
        raise ValueError(
            f"Unrecognised file format: {filepath.name}\n"
            "  Supported: .xlsx (annual XLSX) or .xls (HTML-disguised monthly)."
        )

    # Patch the series field if this is an export file (annual parser always
    # emits SERIES_VALUE; monthly parser does the same).
    if series != SERIES_VALUE:
        for row in rows:
            row["series"] = series

    return rows


# ---------------------------------------------------------------------------
# Google Sheets helpers
# ---------------------------------------------------------------------------

def _load_config() -> dict:
    """Read config.yaml from repo root. Returns empty dict on failure."""
    if not CONFIG_PATH.exists():
        return {}
    try:
        with CONFIG_PATH.open("r", encoding="utf-8") as fh:
            return yaml.safe_load(fh) or {}
    except yaml.YAMLError:
        return {}


def connect_sheets(sheet_id: str):
    """
    Connect to Google Sheets using service account credentials.

    L-3: GOOGLE_SERVICE_ACCOUNT_JSON must be single-line JSON.
    Returns gspread.Spreadsheet object. Calls sys.exit(1) on failure.
    """
    load_dotenv(REPO_ROOT / ".env")

    sa_json_raw = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_json_raw:
        print(
            "ERROR: GOOGLE_SERVICE_ACCOUNT_JSON environment variable is not set.\n"
            "  Local dev: add it to .env at the repo root (single-line JSON — L-3).\n"
            "  GitHub Actions: add it to repository Secrets.",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        sa_info = json.loads(sa_json_raw)
    except json.JSONDecodeError as exc:
        print(
            f"ERROR: GOOGLE_SERVICE_ACCOUNT_JSON is not valid JSON — {exc}\n"
            "  Minify: python -c \"import json,sys; print(json.dumps(json.load(sys.stdin),"
            " separators=(',',':')))\" < key.json",
            file=sys.stderr,
        )
        sys.exit(1)

    creds = Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    gc = gspread.authorize(creds)

    try:
        spreadsheet = gc.open_by_key(sheet_id)
    except gspread.exceptions.APIError as exc:
        print(
            f"ERROR: Could not open sheet {sheet_id} — {exc}\n"
            "  Check the service account has Editor access to the sheet.",
            file=sys.stderr,
        )
        sys.exit(1)

    return spreadsheet


def resolve_sheet_id() -> str:
    """
    Resolve the Google Sheet ID from environment then config.yaml fallback.

    Priority: VKH_SHEET_ID env var → config.yaml sheet_id.
    Calls sys.exit(1) if neither is set.
    """
    load_dotenv(REPO_ROOT / ".env")

    sheet_id = os.environ.get("VKH_SHEET_ID", "").strip()
    if sheet_id:
        return sheet_id

    config = _load_config()
    sheet_id = config.get("sheet_id", "").strip()
    if sheet_id:
        print("  (VKH_SHEET_ID not set — using sheet_id from config.yaml)")
        return sheet_id

    print(
        "ERROR: Sheet ID not found.\n"
        "  Set VKH_SHEET_ID in .env, or ensure sheet_id is set in config.yaml.",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Dedup and write helpers
# ---------------------------------------------------------------------------

def _normalise_hs_code(raw) -> str:
    """
    Normalise an hs_code value to a canonical dot-notation string.

    Google Sheets returns numeric cells as float (e.g. 507.9) even when the
    stored value is the string "0507.90". Converting via str() produces "507.9",
    which does not match the parser output "0507.90" — causing silent dedup
    failure (L-9). This function maps both representations to "0507.90".

    Mapping: 507.9 → "0507.90", 510.0 → "0510.00", "0507.90" → "0507.90".
    Unknown values are returned as-is (str).
    """
    _FLOAT_TO_DOT: dict[float, str] = {
        507.9:  "0507.90",
        510.0:  "0510.00",
    }
    if isinstance(raw, float):
        return _FLOAT_TO_DOT.get(raw, str(raw))
    if isinstance(raw, int):
        return _FLOAT_TO_DOT.get(float(raw), str(raw))
    return str(raw)


def build_dedup_key(row: dict) -> tuple:
    """
    Return the dedup key tuple for a VTW_Trade_Monthly row.

    L-10: key is (date, series, hs_code, unit, country) — five fields.
    Unit distinguishes shipments rows from KG rows.
    L-9: hs_code is normalised via _normalise_hs_code() to handle the float/string
    mismatch between Sheets (returns 507.9) and parser output ("0507.90").
    """
    return (
        str(row.get("date", "")),
        str(row.get("series", "")),
        _normalise_hs_code(row.get("hs_code", "")),
        str(row.get("unit", "")),
        str(row.get("country", "")),
    )


def load_existing_keys(worksheet, series_filter: str = SERIES_VALUE) -> tuple[set, int]:
    """
    Read all rows from the worksheet once and return a set of dedup keys.

    L-4: get_all_records() is called exactly once — never inside a loop.
    Filters to the specified series to keep the key set small and correctly
    scoped (import vs export rows use different series values — GAP-2 fix).
    """
    existing_rows = worksheet.get_all_records()
    series_rows = [r for r in existing_rows if r.get("series") == series_filter]
    return {build_dedup_key(r) for r in series_rows}, len(existing_rows)


def rows_to_append(
    new_rows: list[dict],
    existing_keys: set,
    headers: list[str],
) -> tuple[list[list], int]:
    """
    Filter new_rows to those not already in existing_keys.

    Returns (list_of_lists_for_gspread, skipped_count).
    """
    to_write: list[list] = []
    skipped = 0

    for row in new_rows:
        key = build_dedup_key(row)
        if key in existing_keys:
            skipped += 1
            continue
        to_write.append([row.get(h, "") for h in headers])

    return to_write, skipped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Ingest a QIA deer velvet quarantine file into the VTW_Trade_Monthly "
            "tab of VKH_Data Google Sheet. Auto-detects annual XLSX or monthly HTML-XLS."
        )
    )
    parser.add_argument(
        "--file",
        required=True,
        help=(
            "Path to QIA file. Annual: .xlsx (openpyxl). "
            "Monthly: .xls (HTML-disguised, UTF-8)."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and deduplicate without writing to Google Sheets.",
    )
    args = parser.parse_args()

    file_path = Path(args.file).resolve()

    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    # Detect format for user-facing message.
    if file_path.suffix.lower() == ".xlsx":
        fmt = "annual XLSX"
    elif _is_html_xls(file_path):
        fmt = "monthly HTML-XLS"
    else:
        fmt = "unknown"

    # Derive series before parsing so it can be reported to the user.
    series_for_file = _derive_series(file_path)

    print("ingest_qia.py — VKH QIA quarantine ingestion")
    print(f"  file: {file_path.name}")
    print(f"  format: {fmt}")
    print(f"  series: {series_for_file}")
    print(f"  dry-run: {args.dry_run}")

    # --- Parse ----------------------------------------------------------------
    try:
        parsed_rows = parse_qia_file(file_path)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Unexpected parse failure — {exc}", file=sys.stderr)
        sys.exit(1)

    rows_parsed = len(parsed_rows)

    if rows_parsed == 0:
        print("  WARNING: no rows parsed. Check file format and content.")
        print(f"rows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: 0 | errors: 0")
        sys.exit(0)

    if args.dry_run:
        print()
        print("[DRY RUN] Parsing complete — no Sheets write.")
        print(f"  rows_parsed: {rows_parsed}")
        print("  Sample rows (first 3):")
        for row in parsed_rows[:3]:
            print(f"    {row}")
        print(f"\nrows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: 0 | errors: 0")
        sys.exit(0)

    # --- Connect to Sheets ---------------------------------------------------
    sheet_id = resolve_sheet_id()
    print(f"  sheet_id: {sheet_id}")

    spreadsheet = connect_sheets(sheet_id)
    print(f"  sheet title: {spreadsheet.title}")

    try:
        ws = spreadsheet.worksheet(TARGET_TAB)
    except gspread.exceptions.WorksheetNotFound:
        print(
            f"ERROR: tab '{TARGET_TAB}' not found in sheet {sheet_id}.\n"
            "  Run scripts/setup_sheets.py first to create all Phase 1 tabs.",
            file=sys.stderr,
        )
        sys.exit(1)

    # --- Dedup (L-4: one API call to read all existing rows) -----------------
    existing_keys, existing_count = load_existing_keys(ws, series_filter=series_for_file)
    print(f"  existing rows in tab: {existing_count}")

    headers = ws.row_values(1)
    if not headers:
        print(
            f"ERROR: tab '{TARGET_TAB}' has no header row. "
            "Re-run setup_sheets.py to restore it.",
            file=sys.stderr,
        )
        sys.exit(1)

    new_rows_lists, rows_skipped = rows_to_append(parsed_rows, existing_keys, headers)
    rows_new = len(new_rows_lists)
    errors = 0

    if rows_new == 0:
        print("  Nothing to write — all rows already present.")
        print(f"\nrows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: {rows_skipped} | errors: {errors}")
        sys.exit(0)

    # --- Write (L-4: batch append — never loop) ------------------------------
    rows_written = 0
    batch_size = 200

    for batch_start in range(0, len(new_rows_lists), batch_size):
        batch = new_rows_lists[batch_start:batch_start + batch_size]
        try:
            ws.append_rows(batch, value_input_option="USER_ENTERED")
            rows_written += len(batch)
            if batch_start + batch_size < len(new_rows_lists):
                time.sleep(_BATCH_SLEEP)
        except Exception as exc:  # noqa: BLE001
            logger.error("Sheets write error on batch starting %d: %s", batch_start, exc)
            errors += 1
            break

    print()
    print(f"rows_parsed: {rows_parsed} | rows_written: {rows_written} | rows_skipped: {rows_skipped} | errors: {errors}")

    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
