# Run as: PYTHONPATH=. python scripts/ingest_vfi_records.py --historical /path/to/file.xlsx
#
# ingest_vfi_records.py — VFI import records ingestion for Velvet Knowledge Hub.
#
# Two modes:
#   --historical <file>  Parse the Commander's historical Excel file
#                        (Korean food imports list for deer velvet.xlsx) and
#                        upsert 576 historical rows to VKH VFI_Import_Records tab.
#                        One-time migration from VFI predecessor project.
#
#   --mfds <file>        Parse a new MFDS portal download
#                        (수입식품조회YYYYMMDD.xlsx) and upsert new rows.
#
# Usage:
#   PYTHONPATH=. python scripts/ingest_vfi_records.py --historical /path/to/file.xlsx
#   PYTHONPATH=. python scripts/ingest_vfi_records.py --mfds /path/to/수입식품조회20260520.xlsx
#   PYTHONPATH=. python scripts/ingest_vfi_records.py --historical /path/to/file.xlsx --dry-run
#
# L-1:  PYTHONPATH=. ensures repo root is importable.
# L-2:  .env must be at repo root (/Users/Qs/C/velvet-knowledge-hub/.env).
# L-3:  GOOGLE_SERVICE_ACCOUNT_JSON must be single-line JSON in .env.
# L-4:  get_all_records() called once; batch writes with 200-row batches / 1.1s sleep.
# L-10: Dedup key is (date, importer_ko, product_name) — three fields.
# L-13: Dynamic header detection — scan by column name, never by index.
#
# Note on EN translation fields: VKH does not ship a translation_table.py.
# MFDS-only EN fields (importer_en, country_origin_en, country_export_en,
# product_type_en) are set to empty string for MFDS rows — VFI's lookup table
# dependency is intentionally excluded from VKH.
#
# Note on VFI schema mapping: VFI uses a 25-col schema with classifier flags
# (type_frozen, type_dried, type_ambiguous, source, report_no, item_no,
# weight_kg, quantity, country_origin_raw). VKH VFI_Import_Records uses a
# 19-col schema. Classifier flags are dropped; MFDS-only fields go into notes.
#
# Security: no credentials or secrets in this file. All secrets from .env only.

import argparse
import datetime
import logging
import re
import sys
import time
from pathlib import Path

import gspread
import openpyxl

# ---------------------------------------------------------------------------
# L-1: ensure repo root is on sys.path.
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.sheets_auth import _load_config, connect_sheets, resolve_sheet_id  # noqa: E402

logging.basicConfig(
    level=logging.WARNING,
    format="%(levelname)s %(name)s — %(message)s",
)
logger = logging.getLogger("ingest_vfi_records")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
TARGET_TAB = "VFI_Import_Records"

# VKH VFI_Import_Records schema (19 columns) — from setup_sheets.py.
VFI_HEADERS = [
    "date",
    "year",
    "month",
    "day",
    "importer_en",
    "product_type_en",
    "country_origin_en",
    "country_export_en",
    "importer_ko",
    "product_name",
    "product_en",
    "product_type_ko",
    "exporter_en",
    "expiry_date",
    "country_origin_ko",
    "country_export_ko",
    "importer",
    "notes",
]

# Historical Excel tab name (confirmed from VFI historical_xlsx.py).
_HISTORICAL_TAB = "import list"

# Historical Schema A — raw header → VKH field mapping.
# L-13: column lookup is by header name, not by index.
_HISTORICAL_COLUMN_MAP: dict[str, str] = {
    "Year":                      "year",
    "Month":                     "month",
    "Day":                       "day",
    "Importer":                  "importer_en",
    "Translation of type":       "product_type_en",
    "Country of origin":         "country_origin_en",
    "Country of export":         "country_export_en",
    "Importer (Korean)":         "importer_ko",
    "Product name (Korean)":     "product_name",   # VKH field = product_name
    "Product name (English)":    "product_en",
    "Product type (Korean)":     "product_type_ko",
    "Exporter (English)":        "exporter_en",
    "Date":                      "date_raw",
    "Expire date start from":    "expiry_date_raw",
    "Country of origin (KR)":    "country_origin_ko",
    "Country of export (KR)":    "country_export_ko",
}

# Required keywords for historical file validation (L-13 explicit failure).
_HISTORICAL_REQUIRED = {"Year", "Importer (Korean)", "Product name (Korean)", "Date"}

# MFDS Schema B — Korean header → VKH field mapping.
# Supports both old format (pre-2026) and new format (2026-06-29+).
# Old format names listed first; new format names follow — last non-empty write wins,
# so if the new format name is present it overwrites the old format's empty result.
_MFDS_COLUMN_MAP: dict[str, str] = {
    "처리일자":     "date_raw",       # both formats
    "수입업체":     "importer_ko",    # both formats
    "제품명(한글)": "product_name",   # both formats
    "제품명(영문)": "product_en",     # both formats
    "수출국":       "country_export_ko",  # both formats
    "제조국":       "country_origin_ko",  # both formats
    # Old format only:
    "신고번호":     "_report_no",
    "품목제조번호": "_item_no",
    "순중량(KG)":  "_weight_kg",
    "수량(갯수)":  "_quantity",
    "원산지":       "_country_origin_raw",
    "품목류":       "product_type_ko",
    "제조사(영문)": "exporter_en",
    "유통기한":     "expiry_date_raw",
    # New format (2026-06-29+) — overwrite if present:
    "품목(유형)":  "product_type_ko",
    "해외제조업소": "exporter_en",
    "소비기한":    "expiry_date_raw",
    "구분":         "_category",
    "냉동전환번호": "_frozen_conv_no",
    "이력추적번호": "_trace_no",
    "원재료":       "_raw_material",
}

# Required keywords for MFDS header detection (L-13).
_MFDS_REQUIRED = {"처리일자", "수입업체", "제품명(한글)"}

# L-4: batch write parameters.
_BATCH_SIZE = 200
_BATCH_SLEEP = 1.1


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _safe_str(value) -> str:
    """Return stripped string, or empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(value) -> str:
    """
    Return ISO-format date string (YYYY-MM-DD), or empty string.

    Handles:
    - datetime.date / datetime.datetime objects → strftime
    - YYYYMMDD compact integer or string (e.g. 20260525 → 2026-05-25)
    - Empty / None / "-" → ""
    """
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s == "-":
        return ""
    # YYYYMMDD compact format (MFDS portal output).
    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return s


def _derive_ymd(date_iso: str) -> tuple[str, str, str]:
    """
    Extract (year, month, day) strings from a YYYY-MM-DD date string.

    Returns ("", "", "") if the string is not in YYYY-MM-DD format.
    """
    if date_iso and len(date_iso) == 10 and date_iso[4] == "-":
        return date_iso[:4], date_iso[5:7], date_iso[8:10]
    return "", "", ""


# ---------------------------------------------------------------------------
# Historical XLSX parser (Schema A)
# ---------------------------------------------------------------------------

def _parse_historical(filepath: Path) -> list[dict]:
    """
    Parse the 'import list' tab of the historical Excel file.

    Tab layout:
      Row 0 (index 0): title row — skip.
      Row 1 (index 1): 16 English column headers.
      Row 2+ (index 2+): data rows; blank rows skipped.

    L-13: required columns validated by name, not by index.
    Raises ValueError listing missing columns if required headers absent.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if _HISTORICAL_TAB not in wb.sheetnames:
        raise ValueError(
            f"Tab '{_HISTORICAL_TAB}' not found in {filepath.name}.\n"
            f"Available tabs: {wb.sheetnames}"
        )

    ws = wb[_HISTORICAL_TAB]
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 2:
        raise ValueError(
            f"'{_HISTORICAL_TAB}' tab has fewer than 2 rows — cannot parse."
        )

    # Row index 0 is the title row; row index 1 is the header row.
    raw_headers = [_safe_str(h) for h in all_rows[1]]

    # L-13: validate required columns by name — explicit failure if missing.
    missing = _HISTORICAL_REQUIRED - set(raw_headers)
    if missing:
        raise ValueError(
            f"Historical file missing required columns: {missing}\n"
            f"Actual headers (row 2): {raw_headers}"
        )

    col_index: dict[str, int] = {h: i for i, h in enumerate(raw_headers) if h}
    results: list[dict] = []

    for raw_row in all_rows[2:]:
        # Skip blank rows.
        if all(v is None or _safe_str(v) == "" for v in raw_row):
            continue

        row: dict = {}

        # Map raw columns → internal VKH field names.
        for raw_col, vkh_field in _HISTORICAL_COLUMN_MAP.items():
            idx = col_index.get(raw_col)
            if idx is None or idx >= len(raw_row):
                row[vkh_field] = ""
            else:
                row[vkh_field] = raw_row[idx]

        # Normalise date fields.
        date_iso = _parse_date(row.pop("date_raw", ""))
        expiry_iso = _parse_date(row.pop("expiry_date_raw", ""))
        row["date"] = date_iso
        row["expiry_date"] = expiry_iso

        # Derive year/month/day from the ISO date if not already present from columns.
        # (Historical file has separate Year/Month/Day columns; prefer those if non-empty.)
        year_raw = _safe_str(row.get("year", ""))
        month_raw = _safe_str(row.get("month", ""))
        day_raw = _safe_str(row.get("day", ""))

        if not year_raw and date_iso:
            y, mo, d = _derive_ymd(date_iso)
            row["year"] = y
            row["month"] = mo
            row["day"] = d
        else:
            # Clean numeric year/month/day (openpyxl returns them as int).
            row["year"] = year_raw.split(".")[0] if year_raw else ""
            row["month"] = month_raw.split(".")[0] if month_raw else ""
            row["day"] = day_raw.split(".")[0] if day_raw else ""

        # Normalise all string fields.
        for field in [
            "importer_en", "product_type_en", "country_origin_en", "country_export_en",
            "importer_ko", "product_name", "product_en", "product_type_ko",
            "exporter_en", "country_origin_ko", "country_export_ko",
        ]:
            row[field] = _safe_str(row.get(field, ""))

        # Consolidated importer field (VKH schema) = importer_ko preferred.
        row["importer"] = row["importer_ko"] or row["importer_en"]

        # Notes empty for historical rows.
        row["notes"] = ""

        results.append(row)

    logger.info(
        "Historical parse complete: %d data rows from '%s'.",
        len(results),
        _HISTORICAL_TAB,
    )
    return results


# ---------------------------------------------------------------------------
# MFDS XLSX parser (Schema B)
# ---------------------------------------------------------------------------

def _find_mfds_header_row(all_rows: list) -> int:
    """
    Auto-detect the header row index within the first 5 rows.

    Looks for the row containing '처리일자' (primary MFDS date column).
    Returns the 0-based index of the header row.

    L-13: does not hardcode row number — handles title rows appearing /
    disappearing across MFDS portal download versions.

    Raises:
        ValueError: If no header row is found in the first 5 rows.
    """
    for i, row in enumerate(all_rows[:5]):
        row_strs = [_safe_str(c) for c in row]
        # Check if the core detection keyword is present.
        if "처리일자" in row_strs:
            # Verify at minimum the other two required keywords.
            if _MFDS_REQUIRED.issubset(set(row_strs)):
                return i
    # Build a helpful error if not found.
    first_rows_preview = [
        [_safe_str(c) for c in r] for r in all_rows[:3]
    ]
    raise ValueError(
        f"Cannot find MFDS column headers in file.\n"
        f"Required: {_MFDS_REQUIRED}\n"
        f"First 3 rows: {first_rows_preview}"
    )


def _parse_mfds(filepath: Path) -> list[dict]:
    """
    Parse an MFDS portal weekly download XLSX (수입식품조회YYYYMMDD.xlsx).

    Header row is auto-detected within the first 5 rows (L-13).
    MFDS-only fields (신고번호, 품목제조번호, 순중량(KG), 수량(갯수), 원산지)
    are stored as JSON in the notes column.

    Raises ValueError listing missing required columns if L-13 validation fails.
    """
    # read_only=False required: openpyxl read_only mode silently drops shared-string
    # cells in some MFDS portal files, returning only the first inline-string cell.
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = _find_mfds_header_row(all_rows)
    raw_headers = [_safe_str(c) for c in all_rows[header_idx]]

    # Only require the 3 core dedup columns — optional columns vary between
    # old (pre-2026) and new (2026-06-29+) MFDS portal download formats.
    missing = _MFDS_REQUIRED - set(raw_headers)
    if missing:
        raise ValueError(
            f"MFDS file is missing required columns: {missing}\n"
            f"Actual headers: {raw_headers}"
        )

    col_index: dict[str, int] = {h: i for i, h in enumerate(raw_headers) if h}
    results: list[dict] = []

    for raw_row in all_rows[header_idx + 1:]:
        # Skip blank rows.
        if all(v is None or _safe_str(v) == "" for v in raw_row):
            continue

        row: dict = {}

        # Extract raw values using the column map.
        for ko_col, vkh_field in _MFDS_COLUMN_MAP.items():
            idx = col_index.get(ko_col)
            if idx is None or idx >= len(raw_row):
                row[vkh_field] = ""
            else:
                row[vkh_field] = raw_row[idx]

        # Normalise date fields.
        date_iso = _parse_date(row.pop("date_raw", ""))
        expiry_iso = _parse_date(row.pop("expiry_date_raw", ""))
        row["date"] = date_iso
        row["expiry_date"] = expiry_iso

        # Derive year/month/day from MFDS date.
        y, mo, d = _derive_ymd(date_iso)
        row["year"] = y
        row["month"] = mo
        row["day"] = d

        # Normalise string fields.
        for field in [
            "importer_ko", "product_name", "product_en", "product_type_ko",
            "exporter_en", "country_export_ko", "country_origin_ko",
        ]:
            row[field] = _safe_str(row.get(field, ""))

        # EN fields — no translation table in VKH; set to empty string.
        row["importer_en"] = ""
        row["product_type_en"] = ""
        row["country_origin_en"] = ""
        row["country_export_en"] = ""

        # Consolidated importer field.
        row["importer"] = row["importer_ko"]

        # MFDS-only fields → notes as comma-separated key=value pairs.
        mfds_extras: list[str] = []
        for private_field, label in [
            ("_report_no",          "신고번호"),
            ("_item_no",            "품목제조번호"),
            ("_weight_kg",          "순중량(KG)"),
            ("_quantity",           "수량(갯수)"),
            ("_country_origin_raw", "원산지"),
        ]:
            val = _safe_str(row.pop(private_field, ""))
            if val:
                mfds_extras.append(f"{label}={val}")

        row["notes"] = ", ".join(mfds_extras) if mfds_extras else ""

        results.append(row)

    logger.info(
        "MFDS parse complete: %d data rows from %s.",
        len(results),
        filepath.name,
    )
    return results


# ---------------------------------------------------------------------------
# Dedup and write helpers
# ---------------------------------------------------------------------------
#
# _load_config, connect_sheets, resolve_sheet_id imported from
# scripts.sheets_auth (see H-3 fix, VKH audit 2026-07-01), matching the
# pattern used in ingest_kstat.py, ingest_qia.py, classify_articles.py, etc.
#
# build_dedup_key/load_existing_keys/rows_to_append stay local (M-1, not
# fixed): this file's dedup key is (date, importer_ko, product_name), a
# genuinely different shape from ingest_common's 5-field trade key, so
# importing the shared function would not serve this file's needs.

def build_dedup_key(row: dict) -> tuple:
    """
    Return the dedup key tuple for a VFI_Import_Records row.

    L-10: key is (date, importer_ko, product_name) — three fields.
    """
    return (
        str(row.get("date", "")),
        str(row.get("importer_ko", "")),
        str(row.get("product_name", "")),
    )


def load_existing_keys(worksheet) -> tuple[set, int]:
    """
    Read all rows from the worksheet once and return a set of dedup keys.

    L-4: get_all_records() is called exactly once — never inside a loop.
    """
    existing_rows = worksheet.get_all_records()
    return {build_dedup_key(r) for r in existing_rows}, len(existing_rows)


def rows_to_append(
    new_rows: list[dict],
    existing_keys: set,
    headers: list[str],
) -> tuple[list[list], int]:
    """
    Filter new_rows to those not already in existing_keys.

    Returns (list_of_lists_for_gspread, skipped_count).
    """
    to_write: list[list] = []
    skipped = 0

    for row in new_rows:
        key = build_dedup_key(row)
        if key in existing_keys:
            skipped += 1
            continue
        to_write.append([row.get(h, "") for h in headers])

    return to_write, skipped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Ingest VFI import records into the VFI_Import_Records tab "
            "of VKH_Data Google Sheet."
        )
    )
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument(
        "--historical",
        metavar="FILE",
        help=(
            "Path to historical Excel file "
            "'Korean food imports list for deer velvet.xlsx'."
        ),
    )
    mode_group.add_argument(
        "--mfds",
        metavar="FILE",
        help="Path to MFDS portal download XLSX (수입식품조회YYYYMMDD.xlsx).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and deduplicate without writing to Google Sheets.",
    )
    args = parser.parse_args()

    if args.historical:
        mode = "historical"
        xlsx_path = Path(args.historical).resolve()
    else:
        mode = "mfds"
        xlsx_path = Path(args.mfds).resolve()

    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"ingest_vfi_records.py — VKH VFI records ingestion ({mode} mode)")
    print(f"  file: {xlsx_path.name}")
    print(f"  dry-run: {args.dry_run}")

    # --- Parse ---------------------------------------------------------------
    errors = 0
    try:
        if mode == "historical":
            parsed_rows = _parse_historical(xlsx_path)
        else:
            parsed_rows = _parse_mfds(xlsx_path)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Unexpected parse failure — {exc}", file=sys.stderr)
        sys.exit(1)

    rows_parsed = len(parsed_rows)

    if rows_parsed == 0:
        print("  WARNING: no rows parsed. Check file format.")
        print(f"\nrows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: 0 | errors: 0")
        sys.exit(0)

    if args.dry_run:
        print()
        print("[DRY RUN] Parsing complete — no Sheets write.")
        print(f"  rows_parsed: {rows_parsed}")
        print("  Sample rows (first 3):")
        for row in parsed_rows[:3]:
            print(f"    {row}")
        print(f"\nrows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: 0 | errors: 0")
        sys.exit(0)

    # --- Connect to Sheets ---------------------------------------------------
    sheet_id = resolve_sheet_id()
    print(f"  sheet_id: {sheet_id}")

    spreadsheet = connect_sheets(sheet_id)
    print(f"  sheet title: {spreadsheet.title}")

    try:
        ws = spreadsheet.worksheet(TARGET_TAB)
    except gspread.exceptions.WorksheetNotFound:
        print(
            f"ERROR: tab '{TARGET_TAB}' not found in sheet {sheet_id}.\n"
            "  Run scripts/setup_sheets.py first to create all Phase 1 tabs.",
            file=sys.stderr,
        )
        sys.exit(1)

    # --- Dedup (L-4: one API call) -------------------------------------------
    existing_keys, existing_count = load_existing_keys(ws)
    print(f"  existing rows in tab: {existing_count}")

    headers = ws.row_values(1)
    if not headers:
        print(
            f"ERROR: tab '{TARGET_TAB}' has no header row. "
            "Re-run setup_sheets.py to restore it.",
            file=sys.stderr,
        )
        sys.exit(1)

    new_rows_lists, rows_skipped = rows_to_append(parsed_rows, existing_keys, headers)
    rows_new = len(new_rows_lists)

    if rows_new == 0:
        print("  Nothing to write — all rows already present.")
        print(f"\nrows_parsed: {rows_parsed} | rows_written: 0 | rows_skipped: {rows_skipped} | errors: {errors}")
        sys.exit(0)

    # --- Write (L-4: 200-row batches with 1.1s sleep) ------------------------
    rows_written = 0

    for batch_start in range(0, len(new_rows_lists), _BATCH_SIZE):
        batch = new_rows_lists[batch_start:batch_start + _BATCH_SIZE]
        try:
            ws.append_rows(batch, value_input_option="USER_ENTERED")
            rows_written += len(batch)
            if batch_start + _BATCH_SIZE < len(new_rows_lists):
                time.sleep(_BATCH_SLEEP)
        except Exception as exc:  # noqa: BLE001
            logger.error("Sheets write error on batch starting %d: %s", batch_start, exc)
            errors += 1
            break

    print()
    print(f"rows_parsed: {rows_parsed} | rows_written: {rows_written} | rows_skipped: {rows_skipped} | errors: {errors}")

    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
